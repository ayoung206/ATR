"""Text embedding and table rendering for ATR's offline index.

:class:`Embedder` wraps a BGE-M3 checkout behind a single ``encode`` call that
returns CLS-pooled vectors as a NumPy array. The vectors are stored unnormalised
because the FAISS indices in :mod:`atr.offline.multiview_index` are inner-product
indices built over these raw embeddings.

:func:`excel_to_markdown` renders one workbook as the pipe-delimited table that
the TEXT and HYBRID prompts consume.
"""
from __future__ import annotations

from typing import List, Sequence, Union

import numpy as np
import torch
from openpyxl import load_workbook
from transformers import AutoModel, AutoTokenizer
import os


def resolve_runtime_device(
    requested_device: Union[str, int, None] = "auto",
    require_cuda: bool = False,
    default_cuda_index: int = 0
) -> torch.device:
    cuda_available = torch.cuda.is_available()
    cuda_count = torch.cuda.device_count()
    cuda_version = torch.version.cuda

    def _cuda_unavailable_msg() -> str:
        return (
            "CUDA is not available in current PyTorch runtime. "
            f"(torch.cuda.is_available()={cuda_available}, device_count={cuda_count}, torch_cuda={cuda_version})"
        )

    if requested_device is None:
        requested_device = "auto"

    if isinstance(requested_device, int):
        if not cuda_available:
            raise RuntimeError(_cuda_unavailable_msg())
        return torch.device(f"cuda:{requested_device}")

    requested = str(requested_device).strip()
    requested_lower = requested.lower()

    if requested_lower in ("", "auto"):
        if cuda_available:
            return torch.device(f"cuda:{default_cuda_index}")
        if require_cuda:
            raise RuntimeError(_cuda_unavailable_msg())
        return torch.device("cpu")

    if requested_lower == "cpu":
        if require_cuda:
            raise RuntimeError("require_cuda=True but requested device is cpu.")
        return torch.device("cpu")

    if requested_lower.isdigit():
        if not cuda_available:
            raise RuntimeError(_cuda_unavailable_msg())
        return torch.device(f"cuda:{requested_lower}")

    if requested_lower.startswith("cuda"):
        if not cuda_available:
            raise RuntimeError(_cuda_unavailable_msg())
        return torch.device(requested_lower)

    raise ValueError(f"Unsupported device: {requested_device}")


class Embedder:
    """BGE-M3 encoder: a list of strings in, one vector per string out."""

    def __init__(
        self,
        model_path: str,
        device: Union[str, int, None] = "auto",
        require_cuda: bool = False,
        max_length: int = 512,
    ) -> None:
        self.model_path = model_path
        self.device = resolve_runtime_device(device, require_cuda=require_cuda)
        self.max_length = max_length
        # Half precision is a large speedup on GPU and unavailable on CPU.
        self.use_fp16 = self.device.type == "cuda" and os.getenv("ATR_FP16", "1") == "1"

        self.tokenizer = AutoTokenizer.from_pretrained(model_path, use_fast=True)
        model = AutoModel.from_pretrained(model_path)
        if self.use_fp16:
            model = model.half()
        self.model = model.to(self.device).eval()

    @torch.no_grad()
    def encode(self, texts: Sequence[str]) -> np.ndarray:
        """Encode ``texts`` into a ``(len(texts), hidden_size)`` array."""
        if not texts:
            return np.empty((0, self.model.config.hidden_size), dtype=np.float32)

        batch = self.tokenizer(
            list(texts),
            padding=True,
            truncation=True,
            max_length=self.max_length,
            return_tensors="pt",
        ).to(self.device)

        hidden = self.model(**batch)[0]
        cls_vectors = hidden[:, 0]                 # BGE-M3 pools on the CLS token
        return cls_vectors.float().cpu().numpy()


def excel_to_markdown(file_path: str) -> str:
    """Render every sheet of a workbook as one pipe-delimited table.

    The first row of each sheet is treated as the header and followed by a
    markdown separator row. Empty cells are dropped rather than rendered as
    blanks, matching how the tables were serialised when the index was built.
    """
    workbook = load_workbook(file_path, data_only=True)
    table_name = os.path.basename(file_path).rsplit(".", 1)[0]

    parts: List[str] = [f"Table name: {table_name}\n"]
    for sheet_name in workbook.sheetnames:
        for row_number, row in enumerate(workbook[sheet_name]):
            cells = [str(cell.value) for cell in row if cell.value is not None]
            parts.append(" | " + " | ".join(cells) + " | \n")
            if row_number == 0:
                parts.append(" | " + " | ".join(["---"] * len(cells)) + " | \n")

    return "".join(parts)
